"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import json
import os
import platform
import shutil
import subprocess
import sys
from contextlib import redirect_stderr, redirect_stdout
from io import StringIO
from pathlib import Path

from office.soffice import get_soffice_env
from openpyxl import load_workbook

MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_DIR_WINDOWS = os.path.join(os.environ.get("APPDATA", ""), "LibreOffice/4/user/basic/Standard")
MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

EXCEL_ERRORS = [
    "#VALUE!",
    "#DIV/0!",
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
]


def find_soffice():
    candidates = []
    if os.environ.get("SOFFICE"):
        candidates.append(os.environ["SOFFICE"])

    found = shutil.which("soffice")
    if found:
        candidates.append(found)

    if platform.system() == "Windows":
        candidates.extend(
            [
                r"C:\Program Files\LibreOffice\program\soffice.exe",
                r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            ]
        )

    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return candidate
    return None


def has_gtimeout():
    try:
        subprocess.run(["gtimeout", "--version"], capture_output=True, timeout=1, check=False)
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def setup_libreoffice_macro(soffice_exe):
    if not soffice_exe:
        return False

    system = platform.system()
    if system == "Darwin":
        macro_dir = os.path.expanduser(MACRO_DIR_MACOS)
    elif system == "Windows":
        macro_dir = MACRO_DIR_WINDOWS
    else:
        macro_dir = os.path.expanduser(MACRO_DIR_LINUX)
    
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text():
        return True

    if not os.path.exists(macro_dir):
        try:
            subprocess.run(
                [soffice_exe, "--headless", "--terminate_after_init"],
                capture_output=True,
                timeout=10,
                env=get_soffice_env(),
            )
        except FileNotFoundError:
            return False
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())
    soffice_exe = find_soffice()

    if not setup_libreoffice_macro(soffice_exe):
        return python_formula_check(
            filename,
            "LibreOffice soffice was not found; formulas were evaluated with the Python formulas package without updating cached Excel values.",
        )

    cmd = [
        soffice_exe,
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin" and has_gtimeout():
        cmd = ["gtimeout", str(timeout)] + cmd

    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, env=get_soffice_env(), timeout=timeout
        )
    except subprocess.TimeoutExpired:
        return {"error": f"Recalculation timed out after {timeout} seconds"}

    if result.returncode != 0 and result.returncode != 124:
        error_msg = result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        return {"error": error_msg}

    try:
        result = scan_workbook_errors(filename)
        result["recalculation_engine"] = "libreoffice"
        return result

    except Exception as e:
        return {"error": str(e)}


def scan_workbook_errors(filename):
    error_details = {err: [] for err in EXCEL_ERRORS}
    total_errors = 0

    wb = load_workbook(filename, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    for err in EXCEL_ERRORS:
                        if err in cell.value:
                            location = f"{sheet_name}!{cell.coordinate}"
                            error_details[err].append(location)
                            total_errors += 1
                            break
    wb.close()

    formula_count = count_formulas(filename)
    result = {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "error_summary": {},
        "total_formulas": formula_count,
    }

    for err_type, locations in error_details.items():
        if locations:
            result["error_summary"][err_type] = {
                "count": len(locations),
                "locations": locations[:20],
            }
    return result


def count_formulas(filename):
    wb = load_workbook(filename, data_only=False)
    formula_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
    wb.close()
    return formula_count


def python_formula_check(filename, warning):
    result = scan_workbook_errors(filename)
    result["recalculation_engine"] = "python-formulas"
    result["warning"] = warning

    try:
        import formulas

        with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
            model = formulas.ExcelModel().loads(filename).finish()
            solution = model.calculate()
        for address, value in solution.items():
            text = repr(value)
            for err in EXCEL_ERRORS:
                if err in text:
                    result["error_summary"].setdefault(err, {"count": 0, "locations": []})
                    result["error_summary"][err]["count"] += 1
                    if len(result["error_summary"][err]["locations"]) < 20:
                        result["error_summary"][err]["locations"].append(str(address))
                    result["total_errors"] += 1
        result["status"] = "success" if result["total_errors"] == 0 else "errors_found"
        result["evaluated_formulas"] = result["total_formulas"]
    except Exception as e:
        result["status"] = "verification_incomplete"
        result["formula_evaluation_error"] = str(e)

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
